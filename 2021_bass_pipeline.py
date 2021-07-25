import pandas as pd
import glob
from openpyxl import load_workbook
import os
import sys
import sorting



# post-processing pipeline for BASS csv files
class Pipeline:
    def __init__(self, folder_name):
        self.df_list = []
        for filename in glob.glob(folder_name + '/*.csv'):
            try:
                df1 = pd.read_csv(filename, names=['key', 'value'], skiprows=3, nrows=15)
                df_temp = pd.read_csv(filename, skiprows=18)

                split_index = df_temp.index[df_temp['Observer'] == 'BehavCode'][0]
                df2 = df_temp.iloc[:split_index, :]
                df2.columns = df2.columns.str.lstrip()
                df3 = df_temp.iloc[split_index + 1:, :].dropna(axis=1)
                df3_header = ['BehavCode', 'Occurrences', 'totalTime']
                df3.columns = df3_header
                self.df_list.append([df1, df2, df3, filename[5:-4]])
            except pd.errors.EmptyDataError:
                print(filename + ' is causing problems and cannot be read. Please check it\'s formatting.')
            except KeyError:
                print(filename + ' is causing problems.')
                continue

    def calc_cum_data(self, empty_value):
        """
        Calculates required values for output dataframe
        :return: calculated values in a dataframe format
         """
        cum_data = pd.DataFrame(columns=self.HEADERS)

        for df_group in self.df_list:

            # df1 data
            cum_row = [df_group[3]]
            # cum_data initial order
            for i in [2, 7, 3, 0, 10, 11, 12, 9, 1, None, 12, 4, 5, 6]:
                if i is None:
                    cum_row.append(None)
                else:
                    cum_row.append(df_group[0]['value'][i])

            # df2 data
            behav_dict = {
                'Inactive': [],
                'Locomotion': [],
                'Pacing': [],
                'Jumping': [],
                'Selfbite': [],
                'Selfdirect': [],
                'Swing/spin/flip': [],
                'Headtoss': [],
                'Rock': [],
                'Salute': [],
                'Feargrimace': [],
                'Scratch': [],
                'Yawn': [],
                'Lipsmack': [],
                'Present': [],
                'Cling': [],
                'Mantleshake': [],
                'Vocal': [],
                'Threat/display': [],
                'Aggression': [],
                'Eat/drink': [],
                'Tactile/explore': [],
                'Social/play': [],
                'Groom': [],
                'SocGroom': [],
                'Sex/self/other': [],
                'Other': []
            }
            unique_behavs = 0
            for behav, dur in zip(df_group[1]['BehavCode'], df_group[1]['Duration']):
                # print(dur)
                # values : frequency, total Duration, Avg. Duration, list of timestamps
                if behav == 'Tactile/expl':
                    behav = 'Tactile/explore'
                try:
                    prev_values = behav_dict[behav]
                except KeyError:
                    print(behav)
                    continue
                new_values = []
                if not prev_values:
                    new_values = [1, dur, dur, [dur]]
                    unique_behavs += 1
                else:
                    new_dur_list = prev_values[3]
                    new_dur_list.append(dur)
                    new_values = [
                        prev_values[0] + 1,
                        self.get_total_duration(new_dur_list),
                        self.get_avg_duration(new_dur_list),
                        new_dur_list
                    ]
                behav_dict[behav] = new_values

            # Calculating Values
            most_freq = ['', 0]
            dur_tup_list = []
            NBC_most_freq = ['', 0]
            NBC_dur_tup_list = []
            total_dur_list = []
            for key in behav_dict:
                if behav_dict[key]:
                    freq = behav_dict[key][0]
                    if freq > most_freq[1]:
                        most_freq = [key, freq]
                    dur_tup_list.append((key, behav_dict[key][1]))

                    if key not in self.NBC_dict:
                        if freq > NBC_most_freq[1]:
                            NBC_most_freq = [key, freq]
                        NBC_dur_tup_list.append((key, behav_dict[key][1]))

                    total_dur_list.append(behav_dict[key][1])
            long_dur = self.get_max_duration(dur_tup_list)
            NBC_long_dur = self.get_max_duration(NBC_dur_tup_list)
            total_dur = self.get_total_duration(total_dur_list)

            if not NBC_long_dur:
                NBC_long_dur = (empty_value, '00:00:00.0')
            if not long_dur:
                long_dur = (empty_value, '00:00:00.0')
            if not total_dur:
                total_dur = '00:00:00.0'

            cum_row[10] = total_dur
            cum_row.extend(
                [
                    unique_behavs,
                    most_freq[0],
                    most_freq[1],
                    long_dur[0],
                    long_dur[1],
                    # other_freq_dur[0],
                    # other_freq_dur[1],
                    NBC_most_freq[0],
                    NBC_most_freq[1],
                    NBC_long_dur[0],
                    NBC_long_dur[1]
                ]
            )

            for key in behav_dict:
                if behav_dict[key]:
                    cum_row.extend([behav_dict[key][0], behav_dict[key][1], behav_dict[key][2]])
                else:
                    cum_row.extend([0, '00:00:00.0', empty_value])

            # Adding notes
            if df_group[0]['value'][14]:
                cum_row.append(df_group[0]['value'][14])
            else:
                cum_row.append(empty_value)

            # Adding SIB
            SIB_str = ''
            if behav_dict['Selfbite']:
                if behav_dict['Selfbite'][0] >= 1:
                    SIB_str = 'SIB'
                else:
                    SIB_str = 'not'
            else:
                SIB_str = 'not'
            cum_row.append(SIB_str)

            # Adding % values
            value_list = []
            cmp_dur_list = ['00:04:30.0', '00:01:15.0', '00:02:30.0', '00.01.15.0']
            for behav, cmp_dur in zip(['Inactive', 'Pacing', 'Pacing', 'Vocal'], cmp_dur_list):
                if behav_dict[behav]:
                    max_dur = self.get_max_duration([(behav, behav_dict[behav][1]), (None, cmp_dur)])
                    if max_dur[0] == behav:
                        value_list.append(behav)
                    else:
                        value_list.append('not')
                else:
                    value_list.append('not')
            cum_row.extend(value_list)

            df_length = len(cum_data)
            cum_data.loc[df_length] = cum_row

        return cum_data

    @staticmethod
    def get_avg_duration(dur_list):
        """
        Calculates the average timestamp of all timestamps in a list
        :param dur_list: list of timestamps
        :return: average timestamp string
        """
        total_sec = 0
        for dur in dur_list:
            try:
                total_sec += (int(dur[3:5]) * 60) + float(dur[6:])
            except ValueError:
                total_sec += (int(dur[3:5]) * 60) + float(dur[7:])
        sec_avg = total_sec / len(dur_list)
        time_min = int(sec_avg / 60)
        time_sec = sec_avg - (time_min * 60)
        return f'00:{time_min:02d}:{time_sec:04.1f}'

    @staticmethod
    def get_total_duration(dur_list):
        """
        Calculates the total timestamp of dur_list
        :param dur_list: list of timestamps
        :return: total duration timestamp string
        """
        total_sec = 0
        for dur in dur_list:
            try:
                total_sec += (int(dur[3:5]) * 60) + float(dur[6:])
            except ValueError:
                total_sec += (int(dur[3:5]) * 60) + float(dur[7:])
        time_min = int(total_sec / 60)
        time_sec = total_sec - (time_min * 60)
        return f'00:{time_min:02d}:{time_sec:04.1f}'

    @staticmethod
    def get_max_duration(dur_tup_list):
        """
        Calculates the max timestamp of dur_list
        :param dur_tup_list: list of tuples as (behavior, timestamps)
        :return: max duration timestamp string
        """
        max_dur = [0, '']
        for dur_tup in dur_tup_list:
            try:
                dur_sec = (int(dur_tup[1][3:5]) * 60) + float(dur_tup[1][6:])
            except ValueError:
                dur_sec = (int(dur_tup[1][3:5]) * 60) + float(dur_tup[1][7:])
            if dur_sec > max_dur[0]:
                max_dur = [dur_sec, dur_tup]
        return max_dur[1]

    @staticmethod
    def create_excel(df, filename):
        """
        Creates new excel sheet and exports data without much formatting
        :param df: dataframe to be exported
        :return: none
        """
        df.to_excel('Output/other_' + filename + '.xlsx', index=False)

    @staticmethod
    def export_to_excel(df, filename):
        """
        Appends a dataframe to an existing excel sheet
        :param df: dataframe to be exported
        :return: none
        """
        append_df_to_excel(filename='Output/' + filename + '.xlsx', df=df, startrow=1, index=False, header=False)

    # Constants
    HEADERS = [
        'Sheetname',
        'ID',
        'Date',
        'Site',
        'Obs',
        'start time',
        'End time',
        'Ses Duration',
        'Ethogram',
        'Study code',
        'Tot Duration',
        'Ses Dur',
        'Obs type',
        'Housing',
        'Room/cage',
        'Unique Beh',
        'Most frequent',
        'Freq',
        'Long Dura',
        'Duration',
        'Hi NBC beh Freq',
        'NBC Freq',
        'Hi NBC Beh Dur',
        'NBC Dur',
        'Inactive Frequency',
        'Inactive Duration',
        'Inactive Ave duration',
        'Loco Frequency',
        'Loco Duration',
        'Loco Ave dur',
        'Pacing Frequency',
        'Pacing Duration',
        'Pacing ave dur',
        'Jumping Freq',
        'Jumping Duration',
        'Jumping ave dur',
        'Selfbite Frequency',
        'Selfbite Duration',
        'Selfbite Ave Dur',
        'Selfdirect Frequency',
        'Selfdirect Duration',
        'Selfdirect Ave Dur',
        'Swing/spin/flip Frequency',
        'Swing/spin/flip Duration',
        'Swing/spin/flip Ave Dur',
        'Headtoss Frequency',
        'Headtoss Duration',
        'Headtoss Ave Dur',
        'Rock Frequency',
        'Rock Duration',
        'Rock Ave Dur',
        'Salute Frequency',
        'Salute Duration',
        'Salute Ave Dur',
        'Feargrimace Frequency',
        'Feargrimace Duration',
        'Feargrimace Ave Dur',
        'Scratch Frequency',
        'Scratch Duration',
        'Scratch Ave Dur',
        'Yawn Frequency',
        'Yawn Duration',
        'Yawn Ave Dur',
        'Lipsmack Frequency',
        'Lipsmack Duration',
        'Lipsmack Ave Dur',
        'Present Frequency',
        'Present Duration',
        'Present Ave Dur',
        'Cling Frequency',
        'Cling Duration',
        'Cling Ave Dur',
        'Mantleshake Frequency',
        'Mantleshake Duration',
        'Mantleshake Ave Dur',
        'Vocal Frequency',
        'Vocal Duration',
        'Vocal Ave Dur',
        'Threat/display Frequency',
        'Threat/display Duration',
        'Threat/display Ave Dur',
        'Aggression Frequency',
        'Aggression Duration',
        'Aggression Ave Dur',
        'Eat/drink Frequency',
        'Eat/drink Duration',
        'Eat/drink Ave Dur',
        'Tactile/explore Frequency',
        'Tactile/explore Duration',
        'Tactile/explore Ave Dur',
        'Social/play Frequency',
        'Social/play Duration',
        'Social/play Ave Dur',
        'Groom Frequency',
        'Groom Duration',
        'Groom Ave Dur',
        'SocGroom Frequency',
        'SocGroom Duration',
        'SocGroom Ave Dur',
        'Sex/self/other Frequency',
        'Sex/self/other Duration',
        'Sex/self/other Ave Dur',
        'Other Frequency',
        'Other Duration',
        'Other Ave Dur',
        'Notes',
        'SIB',
        'Inactive 90%',
        'Pacing 25%',
        'Pacing 50%',
        'Vocal 25%'
    ]

    NBC_dict = [
        'Feargrimace',
        'Scratch',
        'Yawn',
        'Lipsmack',
        'Present',
        'Cling',
        'Mantleshake',
        'Vocal',
        'Threat / display',
        'Aggression',
        'Eat / drink',
        'Tactile / explore',
        'Social / play',
        'Groom',
        'SocGroom',
        'Sex / self / other',
        'Other'
    ]


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


if __name__ == '__main__':
    args = sys.argv
    if len(args) == 1:
        args.append(-99)
        args.append('cummulative_data')
    elif len(args) == 2:
        args.append('cummulative_data')

    try:
        assert len(args) == 3
        assert isinstance(args[1], (int, str))
        assert isinstance(args[2], str)
    except AssertionError:
        print('Argument Error: 2021_bass_pipeline.py only takes 2 arguements!')
        print('Arg1 = null value to fill empty data')
        print('Arg2 = name of output.xlsx')
        exit()

    print('Reading data...   ', end='')
    bass_pipe = Pipeline('Data')
    print('Done')
    print('Calculating new values...   ', end='')
    data = bass_pipe.calc_cum_data(args[1])
    print('Done')
    print('Exporting new data to excel files...   ', end='')
    bass_pipe.export_to_excel(data, args[2])
    bass_pipe.create_excel(data, args[2])
    print('Done')
    print('Quicksorting behavior frequencies...   ', end='')
    sorting.main(data, bass_pipe.HEADERS)
    print('Done')
    # SQL Database
    #print('Exporting data to SQL database... ')
    #connection = mysql_db.connect_to_db()
    #mysql_db.write_to_db(data, connection)
    print('Pipeline complete.')
